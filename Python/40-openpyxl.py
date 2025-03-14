import openpyxl

"""El módulo openpyxl es una librería en Python que permite leer, escribir y modificar archivos de Excel en formato .xlsx 
(formato de archivo de Excel 2007 y versiones posteriores). Es útil para trabajar con hojas de cálculo sin necesidad de tener instalado Excel."""

#Funciones principales de openpyxl:
#Leer archivos de Excel: Puedes abrir un archivo Excel y acceder a las celdas de las hojas de cálculo.
#Modificar archivos de Excel: Puedes cambiar valores de celdas, agregar nuevas filas o columnas, y formatear celdas.
#Escribir archivos de Excel: Puedes crear un archivo Excel nuevo y guardarlo.

#Ejemplo básico de uso:
from openpyxl import Workbook, load_workbook 
# Crear un libro de Excel nuevo 
wb = Workbook() 
ws = wb.active 
ws['A1'] = "Hola, Mundo!" # Escribir en una celda 
# Guardar el archivo 
wb.save("ejemplo.xlsx") # Cargar un archivo existente 
wb2 = load_workbook("ejemplo.xlsx") 
ws2 = wb2.active 
print(ws2['A1'].value) # Leer el valor de una celda 
# Modificar el archivo 
ws2['A1'] = "Nuevo valor" 
wb2.save("ejemplo_modificado.xlsx") 

#Características principales:
#Manejo de hojas de cálculo: Puedes acceder a múltiples hojas dentro de un archivo.
#Operaciones con celdas: Leer, escribir, y modificar el contenido de celdas.
#Formato y estilo: Puedes aplicar estilos, cambiar el color de celdas, ajustar anchos de columnas, entre otros.
#Trabajar con fórmulas: Puedes leer y escribir fórmulas en las celdas.


"""El módulo openpyxl tiene muchas funciones"""

#1. openpyxl.load_workbook() Carga un archivo de Excel existente. Ejemplo:
from openpyxl import load_workbook
wb = load_workbook('archivo.xlsx') 

#2. openpyxl.Workbook() Crea un nuevo libro de trabajo (archivo de Excel). Ejemplo: 
from openpyxl import Workbook 
wb = Workbook() 

#3. openpyxl.utils.get_column_letter() Convierte un índice de columna (número) a una letra de columna correspondiente. Ejemplo: 
from openpyxl.utils import get_column_letter 
print(get_column_letter(1)) # 'A' 
print(get_column_letter(27)) # 'AA' 

#4. openpyxl.utils.column_index_from_string() Convierte una letra de columna (como 'A') a su índice numérico correspondiente. Ejemplo: 
from openpyxl.utils import column_index_from_string 
print(column_index_from_string('A')) # 1 
print(column_index_from_string('AA')) # 27 

#5. openpyxl.Workbook.create_sheet() Crea una nueva hoja en el libro de trabajo. Ejemplo: 
from openpyxl import Workbook 
wb = Workbook() 
wb.create_sheet("NuevaHoja") 

#6. openpyxl.Workbook.remove() Elimina una hoja del libro de trabajo. Ejemplo: 
from openpyxl import Workbook 
wb = Workbook() 
hoja = wb.create_sheet("HojaParaEliminar") 
wb.remove(hoja) 

#7. openpyxl.Workbook.active Obtiene la hoja activa del libro de trabajo. Ejemplo: 
from openpyxl import Workbook 
wb = Workbook()
hoja_activa = wb.active 

#8. openpyxl.worksheet.worksheet.Worksheet.cell() Accede a una celda específica de la hoja de trabajo. Ejemplo: 
from openpyxl import Workbook 
wb = Workbook() 
ws = wb.active 
ws['A1'] = "Hola, Mundo!"
print(ws['A1'].value) # "Hola, Mundo!" 

#9. openpyxl.worksheet.worksheet.Worksheet.append() Añade una nueva fila al final de la hoja de trabajo. Ejemplo:
from openpyxl import Workbook
wb = Workbook() 
ws = wb.active 
ws.append([1, 2, 3]) 

#10. openpyxl.worksheet.worksheet.Worksheet.iter_rows() Itera sobre las filas de una hoja de trabajo. Ejemplo: 
from openpyxl import Workbook 
wb = Workbook() 
ws = wb.active 
ws.append([1, 2, 3]) 
for row in ws.iter_rows(values_only=True):
    print(row) # (1, 2, 3) 

#11. openpyxl.worksheet.worksheet.Worksheet.iter_columns() Itera sobre las columnas de una hoja de trabajo. Ejemplo: 
from openpyxl import Workbook 
wb = Workbook() 
ws = wb.active 
ws.append([1, 2, 3]) 
for col in ws.iter_columns(values_only=True): 
    print(col) # (1, 2, 3) 

#12. openpyxl.worksheet.worksheet.Worksheet.merge_cells() Fusiona un rango de celdas en una hoja de trabajo. Ejemplo:
from openpyxl import Workbook 
wb = Workbook() 
ws = wb.active 
ws.merge_cells('A1:C1') 

#13. openpyxl.worksheet.worksheet.Worksheet.unmerge_cells() Deshace la fusión de un rango de celdas. Ejemplo:
from openpyxl import Workbook
wb = Workbook() 
ws = wb.active 
ws.merge_cells('A1:C1') 
ws.unmerge_cells('A1:C1') 

#14. openpyxl.styles.Font() Aplica estilo de fuente (como negrita, cursiva, tamaño, etc.) a una celda. Ejemplo: 
from openpyxl import Workbook 
from openpyxl.styles import Font 
wb = Workbook() 
ws = wb.active 
ws['A1'] = "Texto en negrita" 
ws['A1'].font = Font(bold=True) 

#15. openpyxl.styles.Alignment() Alinea el contenido de una celda (centrado, a la izquierda, a la derecha, etc.). Ejemplo: 
from openpyxl import Workbook 
from openpyxl.styles import Alignment 
wb = Workbook() 
ws = wb.active 
ws['A1'] = "Texto centrado" 
ws['A1'].alignment = Alignment(horizontal="center", vertical="center") 

#16. openpyxl.styles.Color() Cambia el color de fondo o el texto de una celda. Ejemplo: 
from openpyxl import Workbook 
from openpyxl.styles import PatternFill 
wb = Workbook() 
ws = wb.active 
ws['A1'] = "Celda colorida" 
ws['A1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") 

#17. openpyxl.worksheet.worksheet.Worksheet.protection Protege una hoja o celdas específicas contra cambios. Ejemplo: 
from openpyxl import Workbook 
wb = Workbook() 
ws = wb.active 
ws.protection.sheet = True 

#18. openpyxl.drawing.image.Image() Inserta una imagen en una hoja de Excel. Ejemplo: 
from openpyxl import Workbook
from openpyxl.drawing.image import Image 
wb = Workbook() 
ws = wb.active 
img = Image("imagen.png") 
ws.add_image(img, 'A1') 

#19. openpyxl.chart.BarChart() Crea un gráfico de barras en una hoja de trabajo. Ejemplo: 
from openpyxl import Workbook 
from openpyxl.chart import BarChart, Reference 
wb = Workbook() 
ws = wb.active 
ws.append([1, 2, 3, 4]) 
chart = BarChart() 
data = Reference(ws, min_col=1, min_row=1, max_col=4, max_row=1) 
chart.add_data(data, titles_from_data=True) 
ws.add_chart(chart, "A6") 

#20. openpyxl.utils.formulas.translate_formula() Traduce una fórmula de una celda a otro rango o ubicación. Ejemplo: 
from openpyxl.utils import formulas 
formula = "=SUM(A1:B2)" 
new_formula = formulas.translate_formula(formula, "A1", "B2") 



"""El módulo openpyxl contiene varias clases"""

#1. openpyxl.Workbook
#Representa un libro de trabajo de Excel. Es la clase principal para la creación y manipulación de archivos Excel.
#Métodos comunes:
#create_sheet(): Crea una nueva hoja en el libro.
#remove(): Elimina una hoja del libro.
#active: Obtiene la hoja activa del libro. Ejemplo:
from openpyxl import Workbook 
wb = Workbook() 

#2. openpyxl.worksheet.worksheet.Worksheet
#Representa una hoja de trabajo dentro de un libro de Excel. Permite manipular celdas, filas y columnas de la hoja.
#Métodos comunes:
#cell(): Accede a una celda específica.
#append(): Añade una nueva fila.
#iter_rows(): Itera sobre las filas.
#iter_columns(): Itera sobre las columnas. Ejemplo:
from openpyxl import Workbook 
wb = Workbook() 
ws = wb.active 
ws['A1'] = "Hola" 

#3. openpyxl.styles.Font
#Clase que permite configurar el estilo de la fuente (negrita, cursiva, color, tamaño, etc.) en las celdas. Ejemplo: 
from openpyxl.styles import Font 
font = Font(bold=True, italic=True, size=12) 

#4. openpyxl.styles.Alignment
#Clase para alinear el contenido de las celdas (horizontal, vertical, etc.). Ejemplo: 
from openpyxl.styles import Alignment 
alignment = Alignment(horizontal="center", vertical="center") 

#5. openpyxl.styles.PatternFill
#Clase que define el relleno de una celda, incluyendo colores de fondo y patrones. Ejemplo: 
from openpyxl.styles import PatternFill 
fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") 

#6. openpyxl.drawing.image.Image
#Clase para insertar imágenes en una hoja de Excel. Ejemplo: 
from openpyxl.drawing.image import Image 
img = Image("imagen.png") 

#7. openpyxl.chart.BarChart
#Representa un gráfico de barras dentro de una hoja de trabajo. Ejemplo: 
from openpyxl.chart import BarChart 
chart = BarChart() 

#8. openpyxl.chart.PieChart
#Representa un gráfico circular dentro de una hoja de trabajo. Ejemplo: 
from openpyxl.chart import PieChart 
chart = PieChart() 

#9. openpyxl.chart.Reference
#Clase que se usa para hacer referencia a datos dentro de una hoja para gráficos. Ejemplo: 
from openpyxl.chart import Reference 
data = Reference(ws, min_col=1, min_row=1, max_col=4, max_row=5) 

#10. openpyxl.styles.Border
#Clase que permite configurar los bordes de las celdas (tipo de borde, grosor, etc.). Ejemplo: 
from openpyxl.styles import Border, Side 
side = Side(border_style="thin", color="000000") 
border = Border(top=side, left=side, right=side, bottom=side) 

#11. openpyxl.styles.Color
#Clase para definir colores en las celdas. Ejemplo: 
from openpyxl.styles import Color 
color = Color(rgb="FFFF00") 

#12. openpyxl.worksheet.dimensions.Dimension
#Representa las dimensiones de una fila o columna en términos de tamaño. Ejemplo: 
from openpyxl.worksheet.dimensions import Dimension 
dimension = Dimension() 

#13. openpyxl.worksheet.protection.SheetProtection
#Clase que permite proteger una hoja de Excel contra cambios. Ejemplo: 
from openpyxl.worksheet.protection import SheetProtection 
protection = SheetProtection() 

#14. openpyxl.utils.dataframe.dataframe_to_rows
#Convierte un DataFrame de pandas en filas de una hoja de Excel. Ejemplo: 
from openpyxl.utils.dataframe import dataframe_to_rows 
import pandas as pd 
df = pd.DataFrame({'A': [1, 2], 'B': [3, 4]}) 
for row in dataframe_to_rows(df, index=False, header=True): 
    ws.append(row) 


"""En el módulo openpyxl, las variables"""

#1. openpyxl.utils.ROW_DIMENSION
#Variable que define la altura predeterminada de las filas en una hoja de Excel. Valor: 15 Ejemplo: 
from openpyxl.utils import ROW_DIMENSION 
print(ROW_DIMENSION) # Imprime: 15 

#2. openpyxl.utils.COLUMN_DIMENSION
#Variable que define el ancho predeterminado de las columnas en una hoja de Excel. Valor: 8.43 Ejemplo: 
from openpyxl.utils import COLUMN_DIMENSION 
print(COLUMN_DIMENSION) # Imprime: 8.43 

#3. openpyxl.styles.colors.COLOR_INDEX
#Un índice de colores predefinidos que pueden usarse en las celdas de la hoja de Excel. Ejemplo: 
from openpyxl.styles.colors import COLOR_INDEX 
print(COLOR_INDEX["RED"]) # Imprime el valor asociado al color "RED" 

#4. openpyxl.styles.DEFAULT_FONT
#Definición de la fuente predeterminada utilizada en las celdas. Valor: Se define como una fuente con ciertos atributos por defecto. Ejemplo: 
from openpyxl.styles import DEFAULT_FONT 
print(DEFAULT_FONT) # Muestra la fuente por defecto 

#5. openpyxl.shared.EXCEL2007
#Variable que hace referencia a las configuraciones predeterminadas de formato para archivos Excel 2007 y versiones posteriores. Valor: Contiene información sobre las versiones más recientes de los formatos Excel. Ejemplo: 
from openpyxl.shared import EXCEL2007 
print(EXCEL2007) # Imprime la configuración asociada a Excel 2007 

#6. openpyxl.worksheet.worksheet.MAX_ROWS
#Definición del número máximo de filas permitido en una hoja de cálculo. Valor: 1,048,576 filas (como el límite en Excel). Ejemplo:
from openpyxl.worksheet.worksheet import MAX_ROWS 
print(MAX_ROWS) # Imprime: 1048576 

#7. openpyxl.worksheet.worksheet.MAX_COLUMNS
#Definición del número máximo de columnas permitido en una hoja de cálculo. Valor: 16,384 columnas (como el límite en Excel). Ejemplo: from openpyxl.worksheet.worksheet import MAX_COLUMNS 
print(MAX_COLUMNS) # Imprime: 16384 

#8. openpyxl.styles.Alignment
#Aunque no es estrictamente una "variable", Alignment es un objeto utilizado comúnmente para definir la alineación en las celdas. A menudo se hace referencia de forma indirecta. Ejemplo: from openpyxl.styles import Alignment alignment = Alignment(horizontal="center", vertical="center") 


"""El módulo openpyxl define algunas constantes"""

#1. openpyxl.styles.COLOR_INDEX
#Un índice que contiene colores predefinidos que pueden ser usados en las celdas. Ejemplo: 
from openpyxl.styles.colors import COLOR_INDEX 
print(COLOR_INDEX["RED"]) # Muestra el valor del color rojo 

#2. openpyxl.styles.BORDER_THIN
#Define un borde fino para las celdas. Ejemplo: 
from openpyxl.styles import Border, Side 
thin_border = Border(left=Side(border_style="thin")) 

#3. openpyxl.styles.BORDER_MEDIUM
#Define un borde de grosor medio para las celdas. Ejemplo: 
from openpyxl.styles import Border, Side 
medium_border = Border(left=Side(border_style="medium")) 

#4. openpyxl.styles.FILL_SOLID
#Define el tipo de relleno sólido para las celdas. Ejemplo: 
from openpyxl.styles import PatternFill 
solid_fill = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00") 

#5. openpyxl.styles.FILL_NONE
#Define que no hay relleno en la celda. Ejemplo: 
from openpyxl.styles import PatternFill 
no_fill = PatternFill(fill_type="none") 

#6. openpyxl.styles.Alignment.HORIZONTAL_CENTER
#Establece la alineación horizontal de una celda al centro. Ejemplo: 
from openpyxl.styles import Alignment 
alignment = Alignment(horizontal="center") 

#7. openpyxl.styles.Alignment.VERTICAL_CENTER
#Establece la alineación vertical de una celda al centro. Ejemplo: 
from openpyxl.styles import Alignment 
alignment = Alignment(vertical="center") 

#8. openpyxl.styles.Font.BOLD
#Define el estilo de fuente en negrita. Ejemplo: 
from openpyxl.styles import Font 
bold_font = Font(bold=True) 

#9. openpyxl.styles.Font.ITALIC
#Define el estilo de fuente en cursiva. Ejemplo: 
from openpyxl.styles import Font 
italic_font = Font(italic=True) 

#10. openpyxl.styles.Font.UNDERLINE
#Define el estilo de fuente con subrayado. Ejemplo:
from openpyxl.styles import Font 
underline_font = Font(underline="single") 

#11. openpyxl.utils.FORMAT_PERCENTAGE
#Formato de porcentaje para las celdas. Ejemplo: 
from openpyxl.utils import FORMAT_PERCENTAGE 
print(FORMAT_PERCENTAGE) # Muestra el formato de porcentaje 

#12. openpyxl.utils.FORMAT_DATE
#Formato de fecha para las celdas. Ejemplo: 
from openpyxl.utils import FORMAT_DATE 
print(FORMAT_DATE) # Muestra el formato de fecha 

#13. openpyxl.utils.FORMAT_CURRENCY
#Formato de moneda para las celdas. Ejemplo: 
from openpyxl.utils import FORMAT_CURRENCY 
print(FORMAT_CURRENCY) # Muestra el formato de moneda 

#14. openpyxl.worksheet.worksheet.MAX_ROWS
#Define el número máximo de filas en una hoja de Excel. Valor: 1,048,576 filas (máximo permitido por Excel). Ejemplo: 
from openpyxl.worksheet.worksheet import MAX_ROWS 
print(MAX_ROWS) # Imprime: 1048576 

#15. openpyxl.worksheet.worksheet.MAX_COLUMNS
#Define el número máximo de columnas en una hoja de Excel. Valor: 16,384 columnas (máximo permitido por Excel). Ejemplo: 
from openpyxl.worksheet.worksheet import MAX_COLUMNS
print(MAX_COLUMNS) # Imprime: 16384 


"""El módulo openpyxl define varias excepciones"""

#1. openpyxl.utils.exceptions.InvalidFileException
#Se lanza cuando se intenta cargar un archivo que no es un archivo de Excel válido o tiene un formato incorrecto. Ejemplo: 
from openpyxl.utils.exceptions import InvalidFileException 
try:
    wb = openpyxl.load_workbook('archivo_invalido.xlsx') 
except InvalidFileException: 
    print("El archivo no es válido.") 

#2. openpyxl.utils.exceptions.IllegalCharacterError
#Se lanza cuando se encuentran caracteres ilegales en el archivo de Excel. Ejemplo: 
from openpyxl.utils.exceptions import IllegalCharacterError 
try: # Intentar escribir un carácter ilegal 
    ws['A1'] = "Texto con carácter ilegal \x01" 
except IllegalCharacterError: 
    print("Se detectó un carácter ilegal.") 

#3. openpyxl.reader.excel.InvalidFileException
#Esta excepción se genera si se intenta abrir un archivo de Excel dañado o no compatible. Ejemplo:
from openpyxl.reader.excel import InvalidFileException
try: 
    wb = openpyxl.load_workbook('archivo_daño.xlsx')
except InvalidFileException:
    print("El archivo de Excel está dañado o no es válido.") 

#4. openpyxl.worksheet.dimensions.InvalidDimensionError
#Se lanza cuando se encuentra un error al definir el rango de las celdas en una hoja de cálculo (por ejemplo, cuando se supera el límite de filas o columnas). Ejemplo: 
from openpyxl.worksheet.dimensions import InvalidDimensionError 
try: 
    ws = wb.create_sheet("Nueva Hoja", 1) 
    ws['A1048577'] = "Valor" # Fuera del límite de filas 
except InvalidDimensionError: 
    print("Se intentó acceder a una fila o columna fuera de los límites.") 

#5. openpyxl.exceptions.UnsupportedZipFile
#Se lanza cuando el archivo que se intenta abrir no es un archivo ZIP válido, lo cual es necesario para los archivos de Excel (.xlsx) que están basados en el formato ZIP. Ejemplo: 
from openpyxl.exceptions import UnsupportedZipFile 
try: 
    wb = openpyxl.load_workbook('archivo_no_zip.xlsx') 
except UnsupportedZipFile:
    print("El archivo no es un archivo ZIP válido.") 

#6. openpyxl.styles.exceptions.IllegalCharacterError
#Se lanza cuando un carácter no permitido se encuentra al intentar aplicar un estilo (por ejemplo, en los nombres de hojas o celdas). Ejemplo:
from openpyxl.styles.exceptions import IllegalCharacterError
try: 
    ws['A1'] = "Texto con carácter ilegal \x01" 
except IllegalCharacterError: 
    print("Carácter ilegal encontrado en el estilo.") 

#7. openpyxl.cell.cell.InvalidCellException
#Se lanza cuando se intenta acceder a una celda inválida en una hoja de cálculo. Ejemplo: 
from openpyxl.cell.cell import InvalidCellException
try: 
    ws['ZZ1'] = "Valor" # Columna fuera del rango permitido
except InvalidCellException: 
    print("La celda es inválida.") 

#8. openpyxl.writer.excel.InvalidSheetException
#Se lanza cuando una hoja que se está escribiendo en un archivo de Excel no es válida. Ejemplo:
from openpyxl.writer.excel import InvalidSheetException
try: 
    wb.remove_sheet(wb['HojaInvalida']) # Si la hoja no existe 
except InvalidSheetException: 
    print("La hoja no existe o es inválida.") 


"""El módulo openpyxl tiene varios submódulos"""

#1. openpyxl.styles
#Proporciona herramientas para definir y aplicar estilos a las celdas de una hoja de cálculo, como fuentes, colores, bordes, alineación, etc. Ejemplo:
from openpyxl.styles import Font, Color
from openpyxl import Workbook 
wb = Workbook()
ws = wb.active 
ws['A1'] = "Hola" 
ws['A1'].font = Font(size=14, bold=True, color="FF0000") 
wb.save("archivo_con_estilos.xlsx") 

#2. openpyxl.worksheet
#Contiene clases y funciones para trabajar con hojas de cálculo, como la creación de hojas, acceso a celdas, filas y columnas. Ejemplo: 
from openpyxl import Workbook 
from openpyxl.worksheet import Worksheet
wb = Workbook() 
ws = wb.active 
ws['A1'] = "Datos"
wb.save("hoja_de_trabajo.xlsx") 

#3. openpyxl.utils
#Proporciona utilidades para convertir entre formatos y facilitar tareas comunes, como la conversión de índices de fila y columna a letras de Excel, o el formato de celdas. Ejemplo: 
from openpyxl.utils import get_column_letter 
column_letter = get_column_letter(2) # Devuelve 'B' print(column_letter) 

#4. openpyxl.drawing
#Permite agregar gráficos, imágenes y otros objetos de dibujo a las hojas de cálculo. Ejemplo:
from openpyxl.drawing.image import Image 
from openpyxl import Workbook
wb = Workbook() 
ws = wb.active
img = Image('imagen.png')
ws.add_image(img, 'A1') 
wb.save('archivo_con_imagen.xlsx') 

#5. openpyxl.comments
#Contiene clases para agregar y manejar comentarios en las celdas de una hoja de Excel. Ejemplo: 
from openpyxl.comments import Comment 
from openpyxl import Workbook 
wb = Workbook() 
ws = wb.active
comment = Comment("Comentario", "Autor") 
ws['A1'].comment = comment 
wb.save("archivo_con_comentarios.xlsx") 

#6. openpyxl.chart
#Permite la creación de gráficos dentro de las hojas de cálculo, como gráficos de barras, líneas, etc. Ejemplo:
from openpyxl.chart import BarChart, Reference 
from openpyxl import Workbook 
wb = Workbook() 
ws = wb.active 
data = [ ['Mes', 'Ventas'], ['Enero', 40], ['Febrero', 50], ['Marzo', 60] ] 
for row in data: 
    ws.append(row) 
chart = BarChart() 
data = Reference(ws, min_col=2, min_row=1, max_row=4, max_col=2) 
chart.add_data(data, titles_from_data=True) 
ws.add_chart(chart, "E5") 
wb.save("grafico.xlsx") 

#7. openpyxl.formatting
#Proporciona herramientas para formatear celdas en términos de formato de número, alineación, bordes, colores y fuentes. Ejemplo: 
from openpyxl.styles import Alignment
from openpyxl import Workbook 
wb = Workbook() 
ws = wb.active 
ws['A1'] = "Texto centrado" 
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
wb.save('archivo_con_alineacion.xlsx') 

#8. openpyxl.cell
#Contiene funciones y clases para trabajar con celdas individuales en las hojas de Excel. Ejemplo: 
from openpyxl import Workbook
wb = Workbook() 
ws = wb.active
ws['A1'] = "Valor en A1" 
ws['B2'] = "Valor en B2" 
wb.save('celdas.xlsx') 

#9. openpyxl.writer
#Proporciona funcionalidades para guardar y escribir archivos de Excel, en formato .xlsx. Ejemplo:
from openpyxl import Workbook
wb = Workbook() 
ws = wb.active
ws['A1'] = "Datos guardados"
wb.save('archivo_guardado.xlsx') 

